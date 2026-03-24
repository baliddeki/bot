"""
Trade Logger
=============
Logs every trade to an Excel file (.xlsx) with proper formatting.
Also prints to console.

Columns:
  timestamp, action, entry_price, sl, tp1, tp2, lot_size,
  risk_amount, risk_percent, ob_type, fvg_timeframe,
  status, fill_price, close_price, pnl_pips, pnl_amount,
  balance_before, balance_after, notes
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config


class TradeLogger:
    COLUMNS = [
        "timestamp",
        "action",
        "entry_price",
        "sl",
        "tp1",
        "tp2",
        "lot_size",
        "risk_amount",
        "risk_percent",
        "ob_type",
        "fvg_timeframe",
        "status",
        "fill_price",
        "close_price",
        "pnl_pips",
        "pnl_amount",
        "balance_before",
        "balance_after",
        "notes",
    ]

    def __init__(self, log_dir=None):
        self.log_dir = log_dir or config.LOG_DIRECTORY
        os.makedirs(self.log_dir, exist_ok=True)
        self.filepath = os.path.join(self.log_dir, config.TRADE_LOG_FILE)
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        """Create the Excel file with headers if it doesn't exist."""
        if os.path.exists(self.filepath):
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"

        # Write headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Set column widths
        widths = {
            "A": 20,  # timestamp
            "B": 8,  # action
            "C": 12,  # entry_price
            "D": 12,  # sl
            "E": 12,  # tp1
            "F": 12,  # tp2
            "G": 10,  # lot_size
            "H": 12,  # risk_amount
            "I": 12,  # risk_percent
            "J": 10,  # ob_type
            "K": 14,  # fvg_timeframe
            "L": 12,  # status
            "M": 12,  # fill_price
            "N": 12,  # close_price
            "O": 10,  # pnl_pips
            "P": 12,  # pnl_amount
            "Q": 14,  # balance_before
            "R": 14,  # balance_after
            "S": 30,  # notes
        }
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze top row
        ws.freeze_panes = "A2"

        # Add auto-filter
        ws.auto_filter.ref = f"A1:S1"

        wb.save(self.filepath)
        print(f"Trade log created: {self.filepath}")

    def log_signal(
        self, signal, lot_size, risk_amount, risk_percent, balance, notes=""
    ):
        """
        Log when a signal is generated and order is placed.
        Status = "PENDING" until filled.
        """
        row = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": signal["action"],
            "entry_price": signal["entry"],
            "sl": signal["sl"],
            "tp1": signal["tp1"],
            "tp2": signal["tp2"],
            "lot_size": lot_size,
            "risk_amount": round(risk_amount, 2),
            "risk_percent": round(risk_percent, 2),
            "ob_type": signal["ob"]["type"],
            "fvg_timeframe": signal["fvg"].get("timeframe", ""),
            "status": "PENDING",
            "fill_price": "",
            "close_price": "",
            "pnl_pips": "",
            "pnl_amount": "",
            "balance_before": round(balance, 2),
            "balance_after": "",
            "notes": notes,
        }
        self._append_row(row)
        self._print_trade(row)

    def log_fill(self, ticket, fill_price):
        """Update a pending trade to OPEN when the limit order fills."""
        self._update_last_matching(
            "status",
            "PENDING",
            {
                "status": "OPEN",
                "fill_price": fill_price,
                "notes": f"Filled at {fill_price} (ticket {ticket})",
            },
        )
        print(f"  [LOG] Order filled: ticket={ticket} at {fill_price}")

    def log_close(
        self, ticket, close_price, pnl_pips, pnl_amount, balance_after, reason=""
    ):
        """Log when a trade closes (SL, TP, or manual)."""
        self._update_last_matching(
            "status",
            "OPEN",
            {
                "status": reason or "CLOSED",
                "close_price": close_price,
                "pnl_pips": round(pnl_pips, 1),
                "pnl_amount": round(pnl_amount, 2),
                "balance_after": round(balance_after, 2),
                "notes": f"{reason} | ticket {ticket}",
            },
        )
        print(f"  [LOG] Trade closed: {reason} pnl={pnl_amount:.2f}")

    def log_partial_close(self, ticket, percent, pnl_amount):
        """Log a partial close (TP1 hit)."""
        self._append_note(f"TP1 partial {percent}% closed, pnl=${pnl_amount:.2f}")

    def log_cancel(self, reason=""):
        """Log when a pending order is cancelled."""
        self._update_last_matching(
            "status",
            "PENDING",
            {
                "status": "CANCELLED",
                "notes": reason,
            },
        )
        print(f"  [LOG] Order cancelled: {reason}")

    def log_rejection(self, signal, reason):
        """Log when a signal is rejected (risk limit, etc)."""
        row = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": signal["action"],
            "entry_price": signal["entry"],
            "sl": signal["sl"],
            "tp1": signal["tp1"],
            "tp2": signal["tp2"],
            "lot_size": "",
            "risk_amount": "",
            "risk_percent": "",
            "ob_type": signal["ob"]["type"],
            "fvg_timeframe": signal["fvg"].get("timeframe", ""),
            "status": "REJECTED",
            "fill_price": "",
            "close_price": "",
            "pnl_pips": "",
            "pnl_amount": "",
            "balance_before": "",
            "balance_after": "",
            "notes": reason,
        }
        self._append_row(row)
        print(f"  [LOG] Signal rejected: {reason}")

    # ================================================================
    # INTERNAL HELPERS
    # ================================================================

    def _append_row(self, row_dict):
        """Append a row to the Excel file."""
        wb = load_workbook(self.filepath)
        ws = wb["Trades"]

        values = [row_dict.get(col, "") for col in self.COLUMNS]
        ws.append(values)

        # Color-code the status cell
        row_num = ws.max_row
        status_col = self.COLUMNS.index("status") + 1
        status_cell = ws.cell(row=row_num, column=status_col)
        self._color_status(status_cell)

        # Color PnL
        pnl_col = self.COLUMNS.index("pnl_amount") + 1
        pnl_cell = ws.cell(row=row_num, column=pnl_col)
        if isinstance(pnl_cell.value, (int, float)):
            if pnl_cell.value > 0:
                pnl_cell.font = Font(color="008000")  # Green
            elif pnl_cell.value < 0:
                pnl_cell.font = Font(color="FF0000")  # Red

        wb.save(self.filepath)

    def _update_last_matching(self, match_col, match_val, updates):
        """Find the last row where match_col == match_val, and update it."""
        wb = load_workbook(self.filepath)
        ws = wb["Trades"]

        col_idx = self.COLUMNS.index(match_col) + 1

        # Search from bottom up
        target_row = None
        for row_num in range(ws.max_row, 1, -1):
            cell_val = ws.cell(row=row_num, column=col_idx).value
            if cell_val == match_val:
                target_row = row_num
                break

        if target_row is None:
            wb.close()
            return

        for key, value in updates.items():
            if key in self.COLUMNS:
                update_col = self.COLUMNS.index(key) + 1
                ws.cell(row=target_row, column=update_col, value=value)

        # Re-color status
        status_col = self.COLUMNS.index("status") + 1
        self._color_status(ws.cell(row=target_row, column=status_col))

        wb.save(self.filepath)

    def _append_note(self, note_text):
        """Append text to the notes column of the last row."""
        wb = load_workbook(self.filepath)
        ws = wb["Trades"]
        notes_col = self.COLUMNS.index("notes") + 1
        last_row = ws.max_row
        existing = ws.cell(row=last_row, column=notes_col).value or ""
        ws.cell(row=last_row, column=notes_col, value=f"{existing} | {note_text}")
        wb.save(self.filepath)

    def _color_status(self, cell):
        """Apply color to a status cell."""
        status = str(cell.value or "").upper()
        colors = {
            "PENDING": ("FFF2CC", "000000"),  # Yellow bg
            "OPEN": ("D6E4F0", "000000"),  # Blue bg
            "CLOSED": ("E2EFDA", "000000"),  # Green bg
            "WIN": ("E2EFDA", "008000"),  # Green
            "LOSS": ("FCE4EC", "FF0000"),  # Red
            "CANCELLED": ("F2F2F2", "808080"),  # Grey
            "REJECTED": ("F2F2F2", "808080"),  # Grey
        }
        bg, fg = colors.get(status, ("FFFFFF", "000000"))
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fg)

    def _print_trade(self, row):
        """Print trade to console."""
        print(
            f"\n{'='*50}\n"
            f"SIGNAL: {row['action']} | {row['ob_type']} OB + {row['fvg_timeframe']} FVG\n"
            f"  Entry: {row['entry_price']}  SL: {row['sl']}\n"
            f"  TP1: {row['tp1']}  TP2: {row['tp2']}\n"
            f"  Lot: {row['lot_size']}  Risk: ${row['risk_amount']} ({row['risk_percent']}%)\n"
            f"  Balance: ${row['balance_before']}\n"
            f"{'='*50}"
        )
